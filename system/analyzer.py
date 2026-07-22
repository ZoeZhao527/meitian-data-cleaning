"""
数据分析引擎 - 读取 Excel 并输出结构化数据
"""
import sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from io import StringIO

try:
    import openpyxl
except ImportError:
    print("需要 openpyxl: pip install openpyxl")
    sys.exit(1)



# ── 配置（从 config.json 加载，无文件则用默认值）──
_BASE = Path(__file__).resolve().parent.parent
_CONFIG_PATH = _BASE / "system" / "config.json"

_config = {}
if _CONFIG_PATH.exists():
    import json
    _config = json.loads(_CONFIG_PATH.read_text(encoding="utf-8"))

DETAIL_FILE = str(_BASE / _config.get("detail_file", "明细-439_带到店时间(4.1-6.30).xlsx"))
ORDER_FILE = str(_BASE / _config.get("order_file", "6月到店_匹配活动.xlsx"))

DEFAULT_MAP = {
    "小程序": "小程序", "天猫": "天猫", "小红书": "小红书",
    "抖音": "抖音", "美大": "美大", "商渠": "商渠",
    "推荐官": "推荐官", "其他": "其他", "山姆": "其他",
}
CHANNEL_MAP = _config.get("channel_mapping", DEFAULT_MAP)


def get_config_info():
    """返回当前配置信息，用于前端展示"""
    return {
        "detail_file": _config.get("detail_file", "明细-439_带到店时间(4.1-6.30).xlsx"),
        "order_file": _config.get("order_file", "6月到店_匹配活动.xlsx"),
        "detail_exists": Path(DETAIL_FILE).exists(),
        "order_exists": Path(ORDER_FILE).exists(),
        "channel_mapping": CHANNEL_MAP,
        "config_path": str(_CONFIG_PATH),
    }



# ── 工具 ──

def parse_date(s):
    if isinstance(s, datetime):
        return s
    if isinstance(s, str):
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d"):
            try:
                return datetime.strptime(s.strip()[:10], fmt)
            except ValueError:
                continue
    return None


def week_label(dt):
    d = dt.day
    m = dt.month
    if d <= 7:
        return f"W1（{m}/1-7）"
    elif d <= 14:
        return f"W2（{m}/8-14）"
    elif d <= 21:
        return f"W3（{m}/15-21）"
    elif d <= 28:
        return f"W4（{m}/22-28）"
    else:
        import calendar
        last = calendar.monthrange(dt.year, dt.month)[1]
        return f"W5（{m}/29-{last}）"


def safe_str(v):
    return "" if v is None else str(v).strip()


def get_sheet_rows(path, sheet_name):
    """通用：读取 Excel 某 sheet 全部行，返回 (headers, rows_list)"""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb[sheet_name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, 1):
            row[h] = ws.cell(r, c).value
        rows.append(row)
    wb.close()
    return headers, rows


# ── 加载 ──

def load_acquisition(start=None, end=None):
    """加载到店纳新数据"""
    path = Path(DETAIL_FILE)
    if not path.exists():
        return []
    _, rows = get_sheet_rows(path, "Sheet1")
    if not start and not end:
        return rows
    filtered = []
    for row in rows:
        dv = row.get("纳新日期") or row.get("下单时间")
        dt = parse_date(dv)
        if dt:
            if start and dt < parse_date(start):
                continue
            if end and dt > parse_date(end):
                continue
        filtered.append(row)
    return filtered


def load_orders(start=None, end=None):
    """加载到店订单"""
    path = Path(ORDER_FILE)
    if not path.exists():
        return []
    _, rows = get_sheet_rows(path, "Sheet1")
    filtered = []
    for row in rows:
        if row.get("订单状态") != "已到店":
            continue
        dv = row.get("到店时间或核销时间") or row.get("下单时间")
        dt = parse_date(dv)
        if dt:
            if start and dt < parse_date(start):
                continue
            if end and dt > parse_date(end):
                continue
        filtered.append(row)
    return filtered


def load_detail_orders(start=None, end=None):
    """从明细文件 Sheet1 加载订单数据（包含渠道信息）"""
    path = Path(DETAIL_FILE)
    if not path.exists():
        return []
    _, rows = get_sheet_rows(path, "Sheet1")
    filtered = []
    for row in rows:
        dv = row.get("下单时间")
        dt = parse_date(dv)
        if dt:
            if start and dt < parse_date(start):
                continue
            if end and dt > parse_date(end):
                continue
        filtered.append(row)
    return filtered


# ── 分析 ──

def analyze_acquisition(rows):
    """纳新分析"""
    by_channel = Counter()
    by_week = defaultdict(lambda: Counter())
    weekly_total = Counter()
    for row in rows:
        raw = safe_str(row.get("纳新渠道", ""))
        ch = CHANNEL_MAP.get(raw, "其他")
        dv = row.get("纳新日期") or row.get("下单时间")
        dt = parse_date(dv)
        wl = week_label(dt) if dt else "未知"
        by_channel[ch] += 1
        by_week[ch][wl] += 1
        weekly_total[wl] += 1
    return {
        "total": sum(by_channel.values()),
        "by_channel": dict(by_channel.most_common()),
        "by_week": {k: dict(v) for k, v in by_week.items()},
        "weekly_total": dict(weekly_total),
        "weeks": sorted(weekly_total.keys()),
    }


def analyze_detail_orders(rows):
    """订单渠道分析"""
    weekly = defaultdict(lambda: defaultdict(int))
    weekly_total = Counter()
    for row in rows:
        fc = safe_str(row.get("first_channel_name", ""))
        sc = safe_str(row.get("second_channel_name", ""))
        dv = row.get("下单时间")
        dt = parse_date(dv)
        wl = week_label(dt) if dt else "未知"
        if fc == "sms":
            ch = "短信"
        elif sc == "pyq":
            ch = "朋友圈"
        elif fc == "总部企微" and sc == "1v1":
            ch = "企微1v1"
        elif fc == "直播":
            ch = "直播"
        elif fc == "总部企微":
            ch = "其他企微"
        else:
            ch = "其他"
        weekly[wl][ch] += 1
        weekly_total[wl] += 1

    all_chs = set()
    for w, d in weekly.items():
        all_chs.update(d.keys())
    channels = sorted(all_chs)

    result = {}
    for w in sorted(weekly.keys()):
        result[w] = {ch: weekly[w].get(ch, 0) for ch in channels}
    return {
        "channels": channels,
        "weekly": result,
        "weekly_total": dict(weekly_total),
        "weeks": sorted(weekly.keys()),
    }


def analyze_order_activity(rows):
    """活动维度分析"""
    counter = Counter()
    for row in rows:
        a = safe_str(row.get("活动名称", ""))
        if a:
            counter[a] += 1
    return dict(counter.most_common())


def get_overview(start=None, end=None):
    """获取首页概览数据"""
    acq_rows = load_acquisition(start, end)
    order_rows = load_orders(start, end)
    detail_rows = load_detail_orders(start, end)

    acq = analyze_acquisition(acq_rows)
    orders = analyze_detail_orders(detail_rows)
    activities = analyze_order_activity(order_rows)

    # 纳新前三渠道
    top_channels = list(acq["by_channel"].items())[:3]

    # 到店率（约数）
    total_orders_all = len(detail_rows)

    return {
        "纳新总数": acq["total"],
        "订单总数": total_orders_all,
        "到店订单数": len(order_rows),
        "活动数": len(activities),
        "纳新渠道TOP3": [{"name": k, "count": v} for k, v in top_channels],
        "纳新趋势": {
            "weeks": acq["weeks"],
            "values": [acq["weekly_total"].get(w, 0) for w in acq["weeks"]],
        },
        "订单趋势": {
            "weeks": orders["weeks"],
            "values": [orders["weekly_total"].get(w, 0) for w in orders["weeks"]],
        },
    }


def get_acquisition_detail(start=None, end=None):
    """获取纳新详细数据"""
    rows = load_acquisition(start, end)
    acq = analyze_acquisition(rows)
    return acq


def get_order_detail(start=None, end=None):
    """获取订单详细数据"""
    detail_rows = load_detail_orders(start, end)
    orders = analyze_detail_orders(detail_rows)
    order_rows = load_orders(start, end)
    activities = analyze_order_activity(order_rows)
    return {"orders": orders, "activities": activities}


def generate_report_md(start=None, end=None, month_label=None):
    """生成 Markdown 报告"""
    acq_rows = load_acquisition(start, end)
    order_rows = load_orders(start, end)
    detail_rows = load_detail_orders(start, end)

    acq = analyze_acquisition(acq_rows)
    orders = analyze_detail_orders(detail_rows)
    activities = analyze_order_activity(order_rows)

    if not month_label:
        dt = parse_date(start or "")
        month_label = dt.strftime("%Y-%m") if dt else "报告"

    buf = StringIO()

    def p(v, t):
        return "0%" if t == 0 else f"{v / t * 100:.0f}%"

    # ── 标题 ──
    buf.write(f"# 运营数据汇报（{month_label}）\n\n")
    buf.write("---\n\n")
    buf.write(f"统计周期：{start} ~ {end}\n\n")

    # ── 一、纳新 ──
    buf.write("## 一、纳新数据汇总\n\n")
    buf.write(f"截至统计周期，纳新总回溯 **{acq['total']}** 个，各渠道贡献如下：\n\n")
    buf.write("| 渠道 | 回溯数 | 占比 |\n|------|-------|------|\n")
    for ch, c in acq["by_channel"].items():
        buf.write(f"| {ch} | {c} | {p(c, acq['total'])} |\n")
    buf.write("\n")
    if acq["total"] > 0:
        top = list(acq["by_channel"].items())[:3]
        top_sum = sum(c for _, c in top)
        top_names = "、".join(ch for ch, _ in top)
        buf.write(f"**{top_names}**是纳新的三大主力渠道，合计占比 {p(top_sum, acq['total'])}。\n")
        last_ch = list(acq["by_channel"].items())[-1]
        buf.write(f"{last_ch[0]}贡献最弱。\n\n")

    # ── 二、按周 ──
    weeks = acq["weeks"]
    buf.write("## 二、按周回溯变化\n\n")
    if len(weeks) >= 2:
        recent = weeks[-2:]
        chs = list(acq["by_channel"].keys())
        buf.write("| 渠道 | " + " | ".join(recent) + " | 变化 |\n")
        buf.write("|------|" + "|".join(["--------"] * len(recent)) + "|------|\n")
        for ch in chs:
            v = [str(acq["by_week"].get(ch, {}).get(w, 0)) for w in recent]
            d = int(v[1]) - int(v[0])
            ds = f"+{d}" if d > 0 else str(d)
            buf.write(f"| {ch} | " + " | ".join(v) + f" | {ds} |\n")
        ws = [acq["weekly_total"].get(w, 0) for w in recent]
        buf.write("| **合计** | " + " | ".join(str(w) for w in ws) + f" | {ws[1] - ws[0]:+d} |\n\n")
        buf.write("**两处明显变动：**\n\n")
        changes = []
        for ch in chs:
            v1 = int(acq["by_week"].get(ch, {}).get(recent[0], 0))
            v2 = int(acq["by_week"].get(ch, {}).get(recent[1], 0))
            changes.append((ch, v1, v2, v2 - v1))
        changes.sort(key=lambda x: abs(x[3]), reverse=True)
        for ch, v1, v2, diff in changes[:3]:
            if diff > 0:
                buf.write(f"1. **{ch}** 回溯量增长，从 {v1} 增至 {v2}（+{diff}）。\n")
            elif diff < 0:
                buf.write(f"1. **{ch}** 回溯量下降，从 {v1} 降至 {v2}（{diff}）。\n")
        buf.write("\n")
    else:
        buf.write("（数据不足以进行周对比）\n\n")

    # ── 三、订单 ──
    od = orders
    buf.write("## 三、订单数据（按周）\n\n")
    ows = od["weeks"]
    if ows:
        chs = od["channels"]
        buf.write("| 周次 | " + " | ".join(chs) + " | 小计 |\n")
        buf.write("|------|" + "|".join(["------"] * len(chs)) + "|------|\n")
        for w in ows:
            v = [str(od["weekly"][w].get(c, 0)) for c in chs]
            sub = od["weekly_total"].get(w, 0)
            buf.write(f"| {w} | " + " | ".join(v) + f" | {sub} |\n")
        buf.write("\n")
        buf.write("### 订单趋势要点\n\n")
        if len(ows) >= 2:
            for c in chs:
                cv = [od["weekly"][w].get(c, 0) for w in ows]
                peak = max(cv)
                if peak >= 10 and cv[-1] == 0:
                    buf.write(f"- **{c}** 订单急剧萎缩：从峰值 {peak} 降至近零，需排查原因。\n")
                elif peak >= 10:
                    fv, lv = cv[0], cv[-1]
                    if lv < fv * 0.5:
                        buf.write(f"- **{c}** 呈下降趋势：从 {fv} 降至 {lv}。\n")
            tl = [od["weekly_total"].get(w, 0) for w in ows]
            if tl[-1] < tl[0] * 0.5:
                drop = (tl[0] - tl[-1]) / tl[0] * 100
                buf.write(f"- 整体订单量呈下降趋势：从 {tl[0]} 降至 {tl[-1]}，降幅 {drop:.0f}%。\n")
        buf.write("\n")

    # ── 四、小结 ──
    buf.write("## 四、小结\n\n")
    buf.write("1. **纳新端**：...（可在此补充运营分析）\n")
    buf.write("2. **订单端**：...（可在此补充运营分析）\n")

    return buf.getvalue()




def list_data_files():
    """列出 data/ 目录下的可用数据文件"""
    data_dir = _BASE / "system" / "data"
    files = []
    if data_dir.exists():
        for f in sorted(data_dir.iterdir()):
            if f.suffix in (".xlsx", ".xls"):
                files.append({
                    "name": f.name,
                    "size": f.stat().st_size,
                    "modified": datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
                })
    return {
        "current_detail": _config.get("detail_file", ""),
        "current_order": _config.get("order_file", ""),
        "available": files,
        "data_dir": str(data_dir),
    }


# ══════════════════════════════════════════════════════════════
# 表格合并
# ══════════════════════════════════════════════════════════════

def get_sheet_names(file_path):
    """获取 Excel 文件的所有 sheet 名"""
    wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def preview_table(file_path, sheet_name=None):
    """预览表格：表头、行数、样本数据"""
    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    total_rows = ws.max_row - 1  # 减去表头
    # 取前 5 行样本
    sample = []
    for r in range(2, min(7, ws.max_row + 1)):
        row = {}
        for c, h in enumerate(headers, 1):
            row[h] = ws.cell(r, c).value
        sample.append(row)
    wb.close()
    return {
        "sheet": sheet_name,
        "sheets": get_sheet_names(file_path),
        "rows": total_rows,
        "columns": len(headers),
        "headers": headers,
        "sample": sample,
    }


def merge_tables(file_a, file_b, key_a, key_b, join_type="inner",
                 cols_a=None, cols_b=None, sheet_a=None, sheet_b=None):
    """
    合并两个 Excel 表格。
    
    参数:
        file_a, file_b: 文件路径
        key_a, key_b: 匹配键的列名
        join_type: "inner" | "left"
        cols_a, cols_b: 要保留的列（None=全部）
        sheet_a, sheet_b: sheet 名（None=第一个）
    
    返回: {"headers": [...], "rows": [[...], ...], "total": N, "error": "..."}
    """
    wb_a = openpyxl.load_workbook(str(file_a), data_only=True)
    wb_b = openpyxl.load_workbook(str(file_b), data_only=True)

    if sheet_a is None:
        sheet_a = wb_a.sheetnames[0]
    if sheet_b is None:
        sheet_b = wb_b.sheetnames[0]

    ws_a = wb_a[sheet_a]
    ws_b = wb_b[sheet_b]

    headers_a = [ws_a.cell(1, c).value for c in range(1, ws_a.max_column + 1)]
    headers_b = [ws_b.cell(1, c).value for c in range(1, ws_b.max_column + 1)]

    if key_a not in headers_a:
        return {"error": "File A 中未找到列: " + str(key_a)}
    if key_b not in headers_b:
        return {"error": "File B 中未找到列: " + str(key_b)}

    # 确定输出列
    out_cols_a = cols_a if cols_a else headers_a
    out_cols_b = [h for h in (cols_b if cols_b else headers_b) if h != key_b]

    # 校验列
    for h in out_cols_a:
        if h not in headers_a:
            return {"error": "File A 中未找到列: " + str(h)}
    for h in out_cols_b:
        if h not in headers_b:
            return {"error": "File B 中未找到列: " + str(h)}

    key_idx_a = headers_a.index(key_a)
    key_idx_b = headers_b.index(key_b)

    # 构建 File B 的索引（保留全部重复项）
    lookup = {}
    for r in range(2, ws_b.max_row + 1):
        key = ws_b.cell(r, key_idx_b + 1).value
        if key is not None:
            if key not in lookup:
                lookup[key] = []
            lookup[key].append(r)

    result_headers = out_cols_a + out_cols_b
    result_rows = []
    matched_keys = set()

    for r in range(2, ws_a.max_row + 1):
        key = ws_a.cell(r, key_idx_a + 1).value
        if key is None:
            continue
        matched = key in lookup
        if join_type == "inner" and not matched:
            continue

        # 取 File A 行数据
        a_data = []
        for h in out_cols_a:
            ci = headers_a.index(h)
            a_data.append(ws_a.cell(r, ci + 1).value)

        if matched:
            matched_keys.add(key)
            # 遍历 File B 中所有匹配行，每条生成一行结果
            for rb in lookup[key]:
                b_data = []
                for h in out_cols_b:
                    ci = headers_b.index(h)
                    b_data.append(ws_b.cell(rb, ci + 1).value)
                result_rows.append(a_data + b_data)
        else:
            result_rows.append(a_data + [None] * len(out_cols_b))

    wb_a.close()
    wb_b.close()

    return {
        "headers": result_headers,
        "rows": result_rows,
        "total": len(result_rows),
        "match_count": len(matched_keys),
        "file_a_rows": ws_a.max_row - 1,
        "file_b_rows": ws_b.max_row - 1,
    }


# ══════════════════════════════════════════════════════════════
# 单表去重
# ══════════════════════════════════════════════════════════════

def dedup_table(file_path, key_columns, sheet_name=None, keep="first"):
    """
    单表去重，按指定列判断重复，保留全部其他列。
    
    参数:
        file_path: 文件路径
        key_columns: 用于判断重复的列名列表，如 ["客户账号"]
        sheet_name: sheet 名（None=第一个）
        keep: "first"（保留首次出现的行）| "last"（保留最后一次出现的行）
    
    返回: {"headers": [...], "total": N, "removed": N, "rows": [[...], ...]}
    """
    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    for col in key_columns:
        if col not in headers:
            wb.close()
            return {"error": "未找到列: " + str(col)}

    key_indices = [headers.index(col) for col in key_columns]

    # 读取全部行
    all_rows = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        all_rows.append(row)

    wb.close()

    result_rows = []
    removed = 0

    if keep == "first":
        seen = set()
        for row in all_rows:
            key = tuple(row[i] for i in key_indices)
            if key not in seen:
                seen.add(key)
                result_rows.append(row)
            else:
                removed += 1
    elif keep == "last":
        seen = {}
        for row in all_rows:
            key = tuple(row[i] for i in key_indices)
            seen[key] = row  # 后面的覆盖前面的
        removed = len(all_rows) - len(seen)
        # 按原顺序返回（保留首次出现的位置）
        seen_order = {}
        for i, row in enumerate(all_rows):
            key = tuple(row[i] for i in key_indices)
            if key not in seen_order:
                seen_order[key] = i
        result_rows = [None] * len(seen)
        for key, row in seen.items():
            result_rows[seen_order[key]] = row

    return {
        "headers": headers,
        "rows": result_rows,
        "total": len(result_rows),
        "removed": removed,
        "original": len(all_rows),
    }


# ══════════════════════════════════════════════════════════════
# 双表去重：主表删除查重表中出现的数据
# ══════════════════════════════════════════════════════════════

def subtract_table(master_path, lookup_path, key_master, key_lookup,
                   sheet_master=None, sheet_lookup=None):
    """在主表中删除查重表里出现的数据行"""
    wb_m = openpyxl.load_workbook(str(master_path), data_only=True)
    wb_l = openpyxl.load_workbook(str(lookup_path), data_only=True)
    if sheet_master is None:
        sheet_master = wb_m.sheetnames[0]
    if sheet_lookup is None:
        sheet_lookup = wb_l.sheetnames[0]
    ws_m = wb_m[sheet_master]
    ws_l = wb_l[sheet_lookup]
    headers_m = [ws_m.cell(1, c).value for c in range(1, ws_m.max_column + 1)]
    headers_l = [ws_l.cell(1, c).value for c in range(1, ws_l.max_column + 1)]
    if key_master not in headers_m:
        wb_m.close(); wb_l.close()
        return {"error": "主表中未找到列: " + str(key_master)}
    if key_lookup not in headers_l:
        wb_m.close(); wb_l.close()
        return {"error": "查重表中未找到列: " + str(key_lookup)}
    key_idx_m = headers_m.index(key_master)
    key_idx_l = headers_l.index(key_lookup)
    lookup_vals = set()
    for r in range(2, ws_l.max_row + 1):
        v = ws_l.cell(r, key_idx_l + 1).value
        if v is not None:
            lookup_vals.add(v)
    result = []
    removed = 0
    for r in range(2, ws_m.max_row + 1):
        v = ws_m.cell(r, key_idx_m + 1).value
        row = [ws_m.cell(r, c).value for c in range(1, ws_m.max_column + 1)]
        if v in lookup_vals:
            removed += 1
        else:
            result.append(row)
    wb_m.close()
    wb_l.close()
    return {"headers": headers_m, "rows": result, "total": len(result), "removed": removed, "original": ws_m.max_row - 1}
